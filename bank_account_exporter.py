from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from report_page import ReportPage
from utils import normalize_widths


def export_bank_account_excel(pages: list[ReportPage], name: str):
    workbook = Workbook()
    
    primary_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    primary_font = Font(color="FFFFFF", bold=True)
    
    ws = workbook.active

    if(ws is None):
        raise Exception("None sheet")

    ws.title = "Cuentas Bancarias"

    ws.append(["Cod. Empleado", "Cedula", "Nombre Completo", "Cod Banco", "Banco", "N. Cuenta"])
    for cell in ws[ws.max_row]:
        cell.font = primary_font
        cell.fill = primary_fill

    for page in pages:
        for section in page.body.sections:
            emp = section.employee
            bank = emp.bank
            ws.append([
                emp.id,
                emp.identification,
                emp.full_name,
                bank.id,
                bank.name,
                emp.account_number
            ])
    
    normalize_widths(workbook)

    workbook.save(name + ".xlsx")
