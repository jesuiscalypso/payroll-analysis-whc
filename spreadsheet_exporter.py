from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pandas.core.frame import itertools

from employee_section import EmployeeSection
from report_page import ReportPage

from pandas import DataFrame
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from typing import Literal

def export_excel(pages: list[ReportPage], name: str, strategy: Literal['styled', 'unstyled'] = 'styled'):
    companies = set()

    for page in pages:
        companies.add(page.header.company.id)

    if len(companies) > 1:
        raise Exception("Cannot process multi-company reports")

    export_styled_excel(pages, name)

def export_unstyled_excel(pages: list[ReportPage], name: str):
    workbook = Workbook()


    # The report itself, separate from its headers, has some metadata. We'll add it to the first sheet since it's already there

    metadata_sheet = workbook['Sheet']
    metadata_sheet.title = 'Metadatos'

    # A lot of this stuff is stationary so we can just nab it from the first page
    first_page = pages[0]
    main_h = first_page.header

    metadata_sheet.append(['PERIODO', 'INICIO', 'FIN'])
    metadata_sheet.append([main_h.period.descriptor, main_h.period.start, main_h.period.end])
    
    metadata_sheet.append([])
    
    metadata_sheet.append(['COD. EMPRESA', 'EMPRESA'])
    metadata_sheet.append([main_h.company.id, main_h.company.name])

    # We need to group these in a sensible manner, so we'll start with the payroll type

    pages_grouped_by_payroll = itertools.groupby(pages, lambda page: page.header.payroll.id)

    for group in pages_grouped_by_payroll:

        # The payroll type defines the tabs, but we also need to separate the entries by section. So we'll do that using the set

        group_pages = group[1]
        printed_organizational_units = set()

        for page in group_pages: 

            sheet_name = f"{page.header.payroll.id} {page.header.payroll.name}"

            if(sheet_name not in workbook.sheetnames):
                worksheet = workbook.create_sheet(sheet_name)
            else:
                worksheet = workbook[sheet_name]
            header = page.header


            # Print the separation if need be.

            organizational_unit_header_text = f"{header.organizational_unit.id} {header.organizational_unit.name}"

            if organizational_unit_header_text not in printed_organizational_units:
                worksheet.append([organizational_unit_header_text])
                worksheet.append([])
                printed_organizational_units.add(organizational_unit_header_text)

            for section in page.body.sections:
                em = section.employee
                op = section.operations
                worksheet.append(['CODIGO', 'NOMBRE', 'CEDULA', 'COD. POSICION', 'POSICION', 'INGRESO', 'RETIRO'])
                worksheet.append([em.id, em.full_name, em.identification, em.position.id, em.position.name, em.entry_date, em.exit_date])
                worksheet.append(['SITUACION', 'COMENTARIO', 'SALIDA', 'REGRESO',])
                worksheet.append([em.situation.name, em.situation.comment, em.situation.departure_date, em.situation.return_date])
                worksheet.append(['SALARIO', 'COD. BANCO', 'BANCO', 'CUENTA'])
                worksheet.append([em.salary, em.bank.id, em.bank.name, em.account_number])
                worksheet.append([])
                worksheet.append(['OPERACIONES'])
                for r in dataframe_to_rows(op, header=True, index=False):
                    worksheet.append(r)
                # Operations here
                worksheet.append([])

    for ws in workbook:
        for column in ws.columns:
            # Get the column letter (e.g., 'A', 'B')
            column_letter = get_column_letter(column[0].column)
            
            # Calculate the maximum length of the cell values in the column
            max_length = 0
            for cell in column:
                try: # Handle potential non-string/None values
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Set the width of the column dimension. Adding a small buffer (e.g., 2 or *1.2) helps.
            if max_length > 0:
                ws.column_dimensions[column_letter].width = max_length + 2 # Add buffer

    workbook.save(name + ".xlsx") 

# I cannot be assed to do this on short notice so we're using AI for this based on the previous method.
def export_styled_excel(pages: list[ReportPage], name: str):
    workbook = Workbook()

    # --- Global Styles ---
    # Deep Blue: Primary Headers
    primary_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    primary_font = Font(color="FFFFFF", bold=True)
    
    # Grey: Secondary/Sub-headers
    sub_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    sub_font = Font(bold=True)
    
    # Gold/Amber: Organizational Unit separator
    org_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    org_font = Font(bold=True, size=13)

    # Metadata Style
    meta_font = Font(bold=True, size=14, color="1F4E78")
    
    center_aligned = Alignment(horizontal="center")

    # --- 1. Metadata Sheet ---
    metadata_sheet = workbook['Sheet']
    metadata_sheet.title = 'Metadatos'
    
    first_page = pages[0]
    main_h = first_page.header

    # Period Info
    metadata_sheet.append(['PERIODO', 'INICIO', 'FIN'])
    for cell in metadata_sheet[metadata_sheet.max_row]:
        cell.font = primary_font
        cell.fill = primary_fill
    metadata_sheet.append([main_h.period.descriptor, main_h.period.start, main_h.period.end])
    
    metadata_sheet.append([]) # Spacer
    
    # Company Info
    metadata_sheet.append(['COD. EMPRESA', 'EMPRESA'])
    for cell in metadata_sheet[metadata_sheet.max_row]:
        cell.font = primary_font
        cell.fill = primary_fill
    metadata_sheet.append([main_h.company.id, main_h.company.name])

    # --- 2. Processing Payroll Groups ---
    pages_grouped_by_payroll = itertools.groupby(pages, lambda page: page.header.payroll.id)

    for group in pages_grouped_by_payroll:
        group_pages = list(group[1])
        printed_organizational_units = set()

        for page in group_pages: 
            sheet_name = f"{page.header.payroll.id} {page.header.payroll.name}"
            
            if sheet_name not in workbook.sheetnames:
                worksheet = workbook.create_sheet(sheet_name)
            else:
                worksheet = workbook[sheet_name]
            
            header = page.header

            # --- Organizational Unit Separator ---
            org_text = f"{header.organizational_unit.id} {header.organizational_unit.name}"
            if org_text not in printed_organizational_units:
                worksheet.append([]) # Top buffer
                worksheet.append([org_text])
                # Style the separator row
                for cell in worksheet[worksheet.max_row]:
                    cell.font = org_font
                    cell.fill = org_fill
                    cell.alignment = center_aligned
                worksheet.append([])
                printed_organizational_units.add(org_text)

            for section in page.body.sections:
                em = section.employee
                op = section.operations
                
                # Employee Header
                worksheet.append(['CODIGO', 'NOMBRE', 'CEDULA', 'COD. POSICION', 'POSICION', 'INGRESO', 'RETIRO'])
                for cell in worksheet[worksheet.max_row]:
                    cell.font = primary_font
                    cell.fill = primary_fill
                
                # Employee Data
                worksheet.append([em.id, em.full_name, em.identification, em.position.id, em.position.name, em.entry_date, em.exit_date])
                
                # Situation & Bank Headers (using sub-style)
                for header_row in [['SITUACION', 'COMENTARIO', 'SALIDA', 'REGRESO'], 
                                   [em.situation.name, em.situation.comment, em.situation.departure_date, em.situation.return_date],
                                   ['SALARIO', 'COD. BANCO', 'BANCO', 'CUENTA'],
                                   [em.salary, em.bank.id, em.bank.name, em.account_number]]:
                    worksheet.append(header_row)
                    # Apply sub-style only to the actual header labels
                    if header_row[0] in ['SITUACION', 'SALARIO']:
                        for cell in worksheet[worksheet.max_row]:
                            cell.font = sub_font
                            cell.fill = sub_fill

                worksheet.append([])
                worksheet.append(['OPERACIONES'])
                worksheet.cell(row=worksheet.max_row, column=1).font = Font(bold=True)

                # Operations Table Logic
                for i, r in enumerate(dataframe_to_rows(op, header=True, index=False)):
                    worksheet.append(r)
                    current_row = worksheet[worksheet.max_row]
                    
                    if i == 0: # DF Header
                        for cell in current_row:
                            cell.font = primary_font
                            cell.fill = primary_fill
                    
                    # Highlight Total Row (checking 'CONCEPTO' column)
                    if current_row[1].value == 'TOTAL POR TRABAJADOR':
                        for cell in current_row:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

                worksheet.append([]) # Bottom spacer for the section

    # --- 3. Final Column Normalization ---
    for ws in workbook:
        for column in ws.columns:
            column_letter = get_column_letter(column[0].column)
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        val_len = len(str(cell.value))
                        if val_len > max_length: max_length = val_len
                except: pass
            
            # Limit width to 60 to prevent extreme cases
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    workbook.save(name + ".xlsx")
